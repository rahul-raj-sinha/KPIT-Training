
from openpyxl import load_workbook

wb = load_workbook("FirstExcel.xlsx")

wb.create_sheet("Players")

wb.active = wb['Players']

ws = wb.active

# Pointing to a cell
cell = ws["A5"]

data = [
        ['PlayerName', "Age", "Runs", "Matches", "Country"],
        ['Sachin', 51, 38650, 950, "India"],
        ["Lara", 52, 24799, 685, "West Indies"],
        ["Pontiong", 48, 32800, 705, "Australia"],
       ["Jayasura", 49, 29450, 689, "Sri lanka"],
    ["Inzamam", 79, 26200, 587, "Palkistan"]
]

for row in data:
    ws.append(row)


wb.save("FirstExcel.xlsx")