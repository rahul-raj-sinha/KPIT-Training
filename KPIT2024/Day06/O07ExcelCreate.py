
from openpyxl import Workbook
from datetime import datetime

wb = Workbook()

ws = wb.active

ws.title = "MyWorkSheet"

ws["B5"] = "Hello World"

ws["B7"] = 1048756

ws["B9"] = datetime.now()

wb.save("FirstExcel.xlsx")
