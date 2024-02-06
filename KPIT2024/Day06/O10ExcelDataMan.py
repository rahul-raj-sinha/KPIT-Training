
from openpyxl import load_workbook

wb = load_workbook("FirstExcel.xlsx")

ws = wb.active

# "A6: E11"

for row in ws.iter_rows(min_row=6, min_col=1, max_row=11):
    for cell in row:
        if cell.value == "Ponting":
            cell.value = "ricky ponting".upper()

wb.save("FirstExcel.xlsx")